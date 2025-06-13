# app_dreux.py – Interface Streamlit Dreux-Gorisse (vibration + Kp + avertissements complets)
# ─────────────────────────────────────────────────────────────────────────────────────────────
import os, sys, importlib.util, builtins, io, contextlib, re
from io import BytesIO

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

# ── 1. Se placer dans le dossier du script

# ── 2. Charger dynamiquement le moteur Dreux
import os
chemin_code = os.path.join(os.path.dirname(__file__), "..", "code_dreux_gorisse_final.py")
spec = importlib.util.spec_from_file_location("dreux", chemin_code)

dreux = importlib.util.module_from_spec(spec)
sys.modules["dreux"] = dreux
spec.loader.exec_module(dreux)

# ── 3. Liste complète de tamis (option 3 du script)
TAMIS = [
    63, 50, 40, 31.5, 25, 20, 16, 12.5, 10, 8, 6.3, 5,
    4, 2.5, 2, 1.25, 1, 0.63, 0.5, 0.315, 0.25, 0.16,
    0.125, 0.08, 0.063
]

# ── 4. Fonction d’exécution (monkey-patch input())
def run_dreux(inputs):
    it = iter(inputs)
    backup = builtins.input
    builtins.input = lambda _="": next(it)
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        dreux.main()
    builtins.input = backup
    return buf.getvalue()

# ── 5. Interface Streamlit
st.set_page_config(page_title="🧱 Dreux-Gorisse", layout="wide")
st.title("🧱 Formulation Béton – Méthode Dreux-Gorisse")

# 5-A. Paramètres béton (sidebar)
st.sidebar.header("Paramètres béton")
fc28 = st.sidebar.number_input("fc28 (MPa)", 10, 100, 25)
qual   = st.sidebar.selectbox("Qualité granulats", ["Excellente", "Bonne", "Passable"], 1)
cim    = st.sidebar.selectbox("Dénomination ciment", ["CEM I 32.5", "CEM I 42.5", "CEM I 52.5"], 0)
aff_cm = st.sidebar.number_input("Affaissement (cm)", 0.0, 20.0, 8.0, 0.5)
type_s = st.sidebar.selectbox("Type sable", ["roulé", "concassé"], 0)
type_g = st.sidebar.selectbox("Type gravier", ["roulé", "concassé"], 1)

st.sidebar.header("Paramètres K′")
vibration = st.sidebar.selectbox("Vibration", ["faible", "normale", "puissante"], 1)
pompabilite = st.sidebar.selectbox("Pompabilité", ["non", "pompable", "très pompable"], 0)

# 5-B. Tableau granulométrique
st.subheader("Granulométrie (2 à 4 granulats)")
n_gr = st.number_input("Nombre de granulats", 2, 4, 3, 1)
g_cols = [f"G{i+1}" for i in range(n_gr)]

df_init = pd.DataFrame(
    {"Tamis (mm)": TAMIS, **{c: [None]*len(TAMIS) for c in g_cols}}
)
granulo_df = st.data_editor(
    df_init, num_rows="fixed", key="granulo", use_container_width=True
)

# 5-C. Masses volumiques
st.subheader("Masses volumiques (kg/m³)")
rho_vals = []
for i, col in enumerate(st.columns(n_gr)):
    default = 2650 if i == 0 else 2600
    rho_vals.append(
        col.number_input(f"ρ {g_cols[i]}", 50, 3000, default, 50)
    )

# ── 6. Bouton Calcul
if st.button("🚀 Lancer le calcul Dreux-Gorisse"):

    # 6-A. Calcul automatique du module de finesse Mf sur G1
    g1_num = pd.to_numeric(granulo_df["G1"], errors="coerce").fillna(0)
    passants = 100 - g1_num
    def pct(t):
        idx = np.where(np.isclose(granulo_df["Tamis (mm)"].astype(float), t, atol=1e-6))[0]
        return float(passants.iloc[idx[0]]) if idx.size else 0.0
    mf_auto = round((pct(2)+pct(1)+pct(0.5)+pct(0.25)+pct(0.125))/100, 2)

    # 6-B. Construire la liste d’inputs pour le script
    inp = ["3", str(n_gr)]
    for g in g_cols:
        inp.append(g)
        inp.extend([str(v) if pd.notna(v) else "" for v in granulo_df[g]])
    inp += [
        str(fc28), qual, cim, str(aff_cm),
        type_s, type_g,
        *map(str, rho_vals),
        type_s, vibration, str(mf_auto), pompabilite
    ]

    # 6-C. Exécution du moteur
    console = run_dreux(inp)

    # 6-D. Extraire tableau de synthèse et avertissements
    synth, warnings_txt = None, ""
    m_table = re.search(r"RÉSULTATS FINAUX\s*:\s*\n(.*)", console, re.S)
    if m_table:
        raw = pd.read_fwf(io.StringIO(m_table.group(1)))
        # repérer la ligne Avertissements
        warn_row = raw[raw.iloc[:,0].str.contains("Avertissements", na=False)]
        if not warn_row.empty:
            warnings_txt = " ".join(
                str(warn_row.iloc[0][c]) for c in raw.columns if pd.notna(warn_row.iloc[0][c])
            ).strip()
            raw = raw.drop(warn_row.index)
        # Nettoyer les colonnes inutiles
        cols = raw.columns
        def last_val(row):
            for c in cols[::-1]:
                val = str(row[c]).strip()
                if val and val.lower() != "nan":
                    return val
            return ""
        synth = pd.DataFrame({
            "Description": raw[cols[0]],
            "Valeur": raw.apply(last_val, axis=1)
        })

    # 6-E. Extraire les coefficients K, Ks, Kp, K′
    m_coef = re.search(r"K=([-\d.]+).*Ks=([-\d.]+).*Kp=([-\d.]+).*K'=\s*([-\d.]+)", console)
    if m_coef:
        K, Ks, Kp, Kprime = map(float, m_coef.groups())
    else:
        K = Ks = Kp = Kprime = None

    # 6-F. Affichage
    st.success(f"✅ Calcul terminé – Mf = {mf_auto} | Vibration : {vibration}, Pompabilité : {pompabilite}")
    if synth is not None:
        st.table(synth)

    if warnings_txt:
        st.warning(f"{warnings_txt}")

    if Kprime is not None:
        met1, met2, met3, met4 = st.columns(4)
        met1.metric("K",     f"{K:.2f}")
        met2.metric("Ks",    f"{Ks:.2f}")
        met3.metric("Kp",    f"{Kp:.2f}")
        met4.metric("K′",    f"{Kprime:.2f}")

    # 6-G. Graphe
    fig = plt.gcf()
    st.pyplot(fig)

    # 6-H. Export PNG
    buf_png = BytesIO()
    fig.savefig(buf_png, format="png", dpi=300, bbox_inches="tight")
    st.download_button("📥 Télécharger courbe (PNG)",
                       buf_png.getvalue(), "courbe_dreux.png", "image/png")

    # 6-I. Export Excel
    wb = Workbook()
    ws_syn = wb.active
    ws_syn.title = "Synthese"
    if synth is not None:
        for r in dataframe_to_rows(synth, index=False, header=True):
            ws_syn.append(r)
        if Kprime is not None:
            ws_syn.append(["K'", Kprime])
    # Feuille Avertissements
    ws_warn = wb.create_sheet("Avertissements")
    if warnings_txt:
        for i, line in enumerate(warnings_txt.splitlines(), start=1):
            ws_warn[f"A{i}"] = line
    else:
        ws_warn["A1"] = "RAS"
    # Feuille Granulo
    ws_gra = wb.create_sheet("Granulo")
    ws_gra.append(["Tamis (mm)", *g_cols])
    for _, row in granulo_df.iterrows():
        ws_gra.append([row["Tamis (mm)"], *row[g_cols].tolist()])
    # Image dans Synthese
    ws_syn.add_image(XLImage(buf_png), "F2")

    buf_xlsx = BytesIO()
    wb.save(buf_xlsx)
    st.download_button("📥 Télécharger synthèse (XLSX)",
                       buf_xlsx.getvalue(), "dreux_synthese.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # 6-J. Console brute
    with st.expander("🔍 Sortie console complète"):
        st.text(console)
